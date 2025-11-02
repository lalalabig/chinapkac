"""
请假管理视图
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Q
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import json

from .models import LeaveApplication, FlightSegment, ApprovalRecord
from .forms import LeaveApplicationForm, ApprovalForm, CancellationForm
from accounts.permissions import role_required
from accounts.models import User


@login_required
def my_applications(request):
    """我的请假申请列表"""
    applications = LeaveApplication.objects.filter(
        applicant=request.user
    ).order_by('-created_at')
    
    # 统计数据
    total_count = applications.count()
    pending_count = applications.filter(
        status__in=[
            LeaveApplication.Status.PENDING_TASK_AREA,
            LeaveApplication.Status.PENDING_HEAD
        ]
    ).count()
    approved_count = applications.filter(
        status=LeaveApplication.Status.APPROVED
    ).count()
    rejected_count = applications.filter(
        status=LeaveApplication.Status.REJECTED
    ).count()
    
    # 分页
    paginator = Paginator(applications, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'applications': page_obj,
        'total_count': total_count,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
    }
    return render(request, 'leave_management/my_applications.html', context)


@login_required
def apply_leave(request):
    """申请请假"""
    if request.method == 'POST':
        try:
            # 解析请假申请数据
            form_data = request.POST.copy()
            
            # 创建请假申请 - 根据申请人角色设置审批流程
            if request.user.role == User.Role.TASK_AREA_MANAGER:
                # 任务区负责人：直接提交到总部负责人
                application = LeaveApplication.objects.create(
                    applicant=request.user,
                    leave_start_date=form_data.get('leave_start_date'),
                    leave_end_date=form_data.get('leave_end_date'),
                    leave_location=form_data.get('leave_location'),
                    leave_latitude=form_data.get('leave_latitude') or None,
                    leave_longitude=form_data.get('leave_longitude') or None,
                    leave_reason=form_data.get('leave_reason'),
                    status=LeaveApplication.Status.PENDING_HEAD
                )
                # 设置总部负责人为当前审批人
                application.current_approver = get_head_manager(request.user)
            else:
                # 普通员工：先提交到任务区负责人
                application = LeaveApplication.objects.create(
                    applicant=request.user,
                    leave_start_date=form_data.get('leave_start_date'),
                    leave_end_date=form_data.get('leave_end_date'),
                    leave_location=form_data.get('leave_location'),
                    leave_latitude=form_data.get('leave_latitude') or None,
                    leave_longitude=form_data.get('leave_longitude') or None,
                    leave_reason=form_data.get('leave_reason'),
                    status=LeaveApplication.Status.PENDING_TASK_AREA
                )
                # 设置任务区负责人为当前审批人
                application.current_approver = get_task_area_manager(request.user)
            application.save()
            
            # 保存回国行程
            outbound_segments = json.loads(form_data.get('outbound_segments', '[]'))
            for idx, segment in enumerate(outbound_segments, 1):
                FlightSegment.objects.create(
                    leave_application=application,
                    segment_type='outbound',
                    sequence=idx,
                    departure=segment.get('departure'),
                    destination=segment.get('destination'),
                    flight_number=segment.get('flight_number'),
                    flight_date=segment.get('flight_date'),
                    flight_time=segment.get('flight_time') or None
                )
            
            # 保存返回行程
            return_segments = json.loads(form_data.get('return_segments', '[]'))
            for idx, segment in enumerate(return_segments, 1):
                FlightSegment.objects.create(
                    leave_application=application,
                    segment_type='return',
                    sequence=idx,
                    departure=segment.get('departure'),
                    destination=segment.get('destination'),
                    flight_number=segment.get('flight_number'),
                    flight_date=segment.get('flight_date'),
                    flight_time=segment.get('flight_time') or None
                )
            
            # 记录审批记录
            ApprovalRecord.objects.create(
                leave_application=application,
                approver=request.user,
                action='submitted',
                comment='提交请假申请'
            )
            
            # 返回JSON响应
            return JsonResponse({
                'success': True,
                'message': '请假申请已提交，等待审批。',
                'redirect_url': '/leave/my-applications/'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'提交失败：{str(e)}'
            }, status=400)
    
    return render(request, 'leave_management/apply_leave.html')


@login_required
def application_detail(request, application_id):
    """请假申请详情"""
    application = get_object_or_404(LeaveApplication, id=application_id)
    
    # 权限检查
    if not can_view_application(request.user, application):
        messages.error(request, '您没有权限查看此申请。')
        return redirect('dashboard:home')
    
    # 获取行程信息
    outbound_flights = application.flight_segments.filter(segment_type='outbound').order_by('sequence')
    return_flights = application.flight_segments.filter(segment_type='return').order_by('sequence')
    
    # 获取审批记录
    approval_records = application.approval_records.all()
    
    context = {
        'application': application,
        'outbound_flights': outbound_flights,
        'return_flights': return_flights,
        'approval_records': approval_records,
        'can_approve': can_approve_application(request.user, application),
        'can_cancel': can_cancel_application(request.user, application),
    }
    return render(request, 'leave_management/application_detail.html', context)


@login_required
@role_required([User.Role.TASK_AREA_MANAGER, User.Role.HEAD_MANAGER, User.Role.SUPERUSER])
def pending_approvals(request):
    """待审批列表"""
    # 根据角色获取待审批列表
    if request.user.role == User.Role.TASK_AREA_MANAGER:
        # 任务区负责人：待任务区审批的申请
        applications = LeaveApplication.objects.filter(
            status=LeaveApplication.Status.PENDING_TASK_AREA,
            applicant__task_area_fk=request.user.task_area_fk
        )
    elif request.user.role == User.Role.HEAD_MANAGER:
        # 总部负责人：待总部审批的申请
        managed_areas = request.user.managed_task_areas.all()
        applications = LeaveApplication.objects.filter(
            status=LeaveApplication.Status.PENDING_HEAD,
            applicant__task_area_fk__in=managed_areas
        )
    else:
        # 超级管理员：所有待审批
        applications = LeaveApplication.objects.filter(
            status__in=[
                LeaveApplication.Status.PENDING_TASK_AREA,
                LeaveApplication.Status.PENDING_HEAD
            ]
        )
    
    applications = applications.order_by('-created_at')
    
    # 分页
    paginator = Paginator(applications, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'applications': page_obj,
    }
    return render(request, 'leave_management/pending_approvals.html', context)


@login_required
@role_required([User.Role.TASK_AREA_MANAGER, User.Role.HEAD_MANAGER, User.Role.SUPERUSER])
def approve_application(request, application_id):
    """审批请假申请"""
    application = get_object_or_404(LeaveApplication, id=application_id)
    
    # 权限检查
    if not can_approve_application(request.user, application):
        messages.error(request, '您没有权限审批此申请。')
        return redirect('leave_management:pending_approvals')
    
    if request.method == 'POST':
        form = ApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            comment = form.cleaned_data['comment']
            
            if action == 'approve':
                # 批准
                if application.status == LeaveApplication.Status.PENDING_TASK_AREA:
                    # 任务区负责人批准
                    application.task_area_manager_approved = True
                    application.task_area_manager_approver = request.user
                    application.task_area_manager_approval_date = timezone.now()
                    
                    # 如果申请人是任务区负责人，直接批准
                    if application.applicant.role == User.Role.TASK_AREA_MANAGER:
                        application.status = LeaveApplication.Status.APPROVED
                        application.current_approver = None
                    else:
                        # 普通员工，流转到总部负责人
                        application.status = LeaveApplication.Status.PENDING_HEAD
                        application.current_approver = get_head_manager(application.applicant)
                    
                    ApprovalRecord.objects.create(
                        leave_application=application,
                        approver=request.user,
                        action='approved',
                        comment=comment or '任务区负责人批准'
                    )
                    messages.success(request, '已批准该请假申请。')
                    
                elif application.status == LeaveApplication.Status.PENDING_HEAD:
                    # 总部负责人批准
                    application.head_manager_approved = True
                    application.head_manager_approver = request.user
                    application.head_manager_approval_date = timezone.now()
                    application.status = LeaveApplication.Status.APPROVED
                    application.current_approver = None
                    
                    ApprovalRecord.objects.create(
                        leave_application=application,
                        approver=request.user,
                        action='approved',
                        comment=comment or '总部负责人批准'
                    )
                    messages.success(request, '已批准该请假申请。')
                
            else:
                # 拒绝
                application.status = LeaveApplication.Status.REJECTED
                application.rejection_reason = comment
                application.current_approver = None
                
                ApprovalRecord.objects.create(
                    leave_application=application,
                    approver=request.user,
                    action='rejected',
                    comment=comment or '审批拒绝'
                )
                messages.info(request, '已拒绝该请假申请。')
            
            application.save()
            return redirect('leave_management:pending_approvals')
    else:
        form = ApprovalForm()
    
    context = {
        'application': application,
        'form': form,
    }
    return render(request, 'leave_management/approve_application.html', context)


@login_required
def cancel_application(request, application_id):
    """取消已批准的请假申请"""
    application = get_object_or_404(LeaveApplication, id=application_id)
    
    # 权限检查
    if not can_cancel_application(request.user, application):
        messages.error(request, '您没有权限取消此申请。')
        return redirect('leave_management:my_applications')
    
    if request.method == 'POST':
        form = CancellationForm(request.POST)
        if form.is_valid():
            cancellation_reason = form.cleaned_data['cancellation_reason']
            
            application.status = LeaveApplication.Status.CANCELLED
            application.cancellation_requested = True
            application.cancellation_reason = cancellation_reason
            application.cancellation_date = timezone.now()
            application.save()
            
            # 记录取消操作
            ApprovalRecord.objects.create(
                leave_application=application,
                approver=request.user,
                action='cancelled',
                comment=cancellation_reason
            )
            
            messages.success(request, '已取消该请假申请。')
            return redirect('leave_management:my_applications')
    else:
        form = CancellationForm()
    
    context = {
        'application': application,
        'form': form,
    }
    return render(request, 'leave_management/cancel_application.html', context)


@login_required
@role_required([User.Role.TASK_AREA_MANAGER, User.Role.HEAD_MANAGER, User.Role.SUPERUSER])
def leave_management_dashboard(request):
    """请假管理仪表板"""
    from accounts.models import TaskArea
    
    # 根据角色获取数据
    if request.user.role == User.Role.TASK_AREA_MANAGER:
        base_query = LeaveApplication.objects.filter(
            applicant__task_area_fk=request.user.task_area_fk
        )
    elif request.user.role == User.Role.HEAD_MANAGER:
        managed_areas = request.user.managed_task_areas.all()
        base_query = LeaveApplication.objects.filter(
            applicant__task_area_fk__in=managed_areas
        )
    else:
        base_query = LeaveApplication.objects.all()
    
    # 获取筛选参数（仅对超级管理员和总部负责人有效）
    task_area_filter = request.GET.get('task_area', '')
    name_filter = request.GET.get('name', '')
    
    # 应用筛选
    if task_area_filter:
        try:
            task_area_filter_id = int(task_area_filter)
            base_query = base_query.filter(applicant__task_area_fk__id=task_area_filter_id)
        except (ValueError, TypeError):
            pass  # 忽略无效的筛选值
    
    if name_filter:
        base_query = base_query.filter(
            Q(applicant__first_name__icontains=name_filter) |
            Q(applicant__last_name__icontains=name_filter) |
            Q(applicant__username__icontains=name_filter)
        )
    
    # 统计数据
    today = timezone.now().date()
    
    # 待审批
    pending_applications = base_query.filter(
        status__in=[
            LeaveApplication.Status.PENDING_TASK_AREA,
            LeaveApplication.Status.PENDING_HEAD
        ]
    ).count()
    
    # 计划休假（已批准但未开始）
    planned_leaves = base_query.filter(
        status=LeaveApplication.Status.APPROVED,
        leave_start_date__gt=today
    )
    
    # 正在休假
    on_leave = base_query.filter(
        status=LeaveApplication.Status.APPROVED,
        leave_start_date__lte=today,
        leave_end_date__gte=today
    )
    
    # 历史记录
    history = base_query.filter(
        status__in=[
            LeaveApplication.Status.APPROVED,
            LeaveApplication.Status.REJECTED,
            LeaveApplication.Status.CANCELLED
        ]
    ).order_by('-created_at')[:10]
    
    # 获取实际有请假申请的任务区（用于筛选下拉框）
    if request.user.role == User.Role.SUPERUSER:
        # 只显示实际有申请的任务区
        task_area_ids = base_query.values('applicant__task_area_fk').distinct()
        task_areas = TaskArea.objects.filter(id__in=task_area_ids)
    elif request.user.role == User.Role.HEAD_MANAGER:
        # 只显示管辖任务区中有申请的
        managed_area_ids = request.user.managed_task_areas.values('id').distinct()
        task_area_ids = base_query.values('applicant__task_area_fk').distinct()
        task_areas = request.user.managed_task_areas.filter(id__in=task_area_ids)
    else:
        task_areas = None
    
    context = {
        'pending_count': pending_applications,
        'planned_leaves': planned_leaves,
        'on_leave': on_leave,
        'history': history,
        'task_areas': task_areas,
        'task_area_filter': task_area_filter,
        'name_filter': name_filter,
    }
    return render(request, 'leave_management/dashboard.html', context)


@login_required
@role_required([User.Role.TASK_AREA_MANAGER, User.Role.HEAD_MANAGER, User.Role.SUPERUSER])
def export_leave_records(request):
    """导出休假记录为Excel"""
    # 根据角色获取数据
    if request.user.role == User.Role.TASK_AREA_MANAGER:
        applications = LeaveApplication.objects.filter(
            applicant__task_area_fk=request.user.task_area_fk
        )
        task_area_name = request.user.task_area_fk.name if request.user.task_area_fk else '未知任务区'
    elif request.user.role == User.Role.HEAD_MANAGER:
        managed_areas = request.user.managed_task_areas.all()
        applications = LeaveApplication.objects.filter(
            applicant__task_area_fk__in=managed_areas
        )
        task_area_name = '所有管辖任务区'
    else:
        applications = LeaveApplication.objects.all()
        task_area_name = '所有任务区'
    
    applications = applications.order_by('-created_at')
    
    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '休假记录'
    
    # 设置标题
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f'{task_area_name}休假情况'
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置表头
    headers = ['姓名', '任务区', '请假开始日期', '请假结束日期', 
               '休假天数', '休假地点', '申请日期', '审批状态', 
               '任务区负责人审批', '总部负责人审批', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 填充数据
    for row_num, app in enumerate(applications, 3):
        ws.cell(row=row_num, column=1).value = f"{app.applicant.last_name}{app.applicant.first_name}"
        ws.cell(row=row_num, column=2).value = app.applicant.task_area_fk.name if app.applicant.task_area_fk else '未设置'
        ws.cell(row=row_num, column=3).value = app.leave_start_date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=4).value = app.leave_end_date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=5).value = app.duration_days
        ws.cell(row=row_num, column=6).value = app.leave_location
        ws.cell(row=row_num, column=7).value = app.application_date.strftime('%Y-%m-%d %H:%M')
        ws.cell(row=row_num, column=8).value = app.get_status_display()
        
        # 审批信息
        if app.task_area_manager_approved:
            approver_name = app.task_area_manager_approver.username if app.task_area_manager_approver else '未知'
            approval_date = app.task_area_manager_approval_date.strftime('%Y-%m-%d') if app.task_area_manager_approval_date else ''
            ws.cell(row=row_num, column=9).value = f"{approver_name} ({approval_date})"
        else:
            ws.cell(row=row_num, column=9).value = '未审批'
        
        if app.head_manager_approved:
            approver_name = app.head_manager_approver.username if app.head_manager_approver else '未知'
            approval_date = app.head_manager_approval_date.strftime('%Y-%m-%d') if app.head_manager_approval_date else ''
            ws.cell(row=row_num, column=10).value = f"{approver_name} ({approval_date})"
        else:
            ws.cell(row=row_num, column=10).value = '未审批'
        
        # 备注
        if app.status == LeaveApplication.Status.REJECTED:
            ws.cell(row=row_num, column=11).value = f"拒绝原因：{app.rejection_reason}"
        elif app.status == LeaveApplication.Status.CANCELLED:
            ws.cell(row=row_num, column=11).value = f"取消原因：{app.cancellation_reason}"
        elif app.is_on_leave:
            ws.cell(row=row_num, column=11).value = '正在休假'
        elif app.is_planned_leave:
            ws.cell(row=row_num, column=11).value = '计划休假'
    
    # 调整列宽
    from openpyxl.cell import MergedCell
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        # 获取列字母（A, B, C等）
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(col_idx)
        
        for cell in col:
            # 跳过合并单元格
            if isinstance(cell, MergedCell):
                continue
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # 设置列宽，最小10，最大50
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 生成响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{task_area_name}休假情况_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# 辅助函数

def get_task_area_manager(user):
    """获取用户的任务区负责人"""
    if not user.task_area_fk:
        return None
    # 使用TaskArea模型中的便捷方法
    return user.task_area_fk.task_area_manager


def get_head_manager(user):
    """获取用户的总部负责人"""
    if not user.task_area_fk:
        return None
    # 查找管理该任务区的总部负责人
    return User.objects.filter(
        managed_task_areas=user.task_area_fk,
        role=User.Role.HEAD_MANAGER
    ).first()


def can_view_application(user, application):
    """检查用户是否可以查看申请"""
    # 申请人自己
    if user == application.applicant:
        return True
    
    # 超级管理员
    if user.role == User.Role.SUPERUSER:
        return True
    
    # 任务区负责人
    if user.role == User.Role.TASK_AREA_MANAGER:
        return application.applicant.task_area_fk == user.task_area_fk
    
    # 总部负责人
    if user.role == User.Role.HEAD_MANAGER:
        managed_areas = user.managed_task_areas.all()
        return application.applicant.task_area_fk in managed_areas
    
    return False


def can_approve_application(user, application):
    """检查用户是否可以审批申请"""
    # 超级管理员可以审批所有
    if user.role == User.Role.SUPERUSER:
        return True
    
    # 任务区负责人
    if user.role == User.Role.TASK_AREA_MANAGER:
        return (
            application.status == LeaveApplication.Status.PENDING_TASK_AREA and
            application.applicant.task_area_fk == user.task_area_fk
        )
    
    # 总部负责人
    if user.role == User.Role.HEAD_MANAGER:
        managed_areas = user.managed_task_areas.all()
        return (
            application.status == LeaveApplication.Status.PENDING_HEAD and
            application.applicant.task_area_fk in managed_areas
        )
    
    return False


def can_cancel_application(user, application):
    """检查用户是否可以取消申请"""
    # 只有申请人自己可以取消
    if user != application.applicant:
        return False
    
    # 只有已批准的申请可以取消
    return application.status == LeaveApplication.Status.APPROVED